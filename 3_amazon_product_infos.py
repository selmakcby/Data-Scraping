import csv
import pandas as pd
from bs4 import BeautifulSoup
from requests_html import HTMLSession
import openpyxl

"""
Extracts product information (title, price, image URL) from each product link in the Excel file
and stores the data in a new Excel file.
"""

# Read URLs from the Excel file into a DataFrame
df_urls = pd.read_excel('unique_product_urls.xlsx')

# Lists to store extracted data
product_titles = []
product_prices = []
product_links = []
image_urls = []

# Start an HTML session
s = HTMLSession()

# Loop through each URL and extract product data
for url in df_urls['Unnamed: 0']:
    # Skip empty or invalid URLs
    if pd.isna(url):
        continue

    # Get and render the page
    r = s.get(url)
    r.html.render(sleep=1)
    
    soup = BeautifulSoup(r.html.html, 'html.parser')

    # Extract product title, price, and image URL
    title = soup.select_one('#productTitle')
    price = soup.find('span', {'class': 'a-offscreen'})
    image = soup.find('img', {'data-a-image-name': 'landingImage'})

    # Store data if all elements are found
    if title and price and image:
        product_titles.append(title.get_text(strip=True))
        product_prices.append(price.get_text(strip=True))
        image_urls.append(image.get('src'))
        product_links.append(url.strip())

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write the headers
sheet['A1'] = 'Title'
sheet['B1'] = 'Price'
sheet['C1'] = 'Link'
sheet['D1'] = 'Image'

# Write the extracted data into the Excel file
for i, (title, price, link, image) in enumerate(zip(product_titles, product_prices, product_links, image_urls), start=2):
    sheet.cell(row=i, column=1).value = title
    sheet.cell(row=i, column=2).value = price
    sheet.cell(row=i, column=3).value = link
    sheet.cell(row=i, column=4).value = image

# Save the workbook
workbook.save('product_info.xlsx')
