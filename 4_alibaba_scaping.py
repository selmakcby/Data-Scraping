from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from bs4 import BeautifulSoup
import time
import openpyxl

"""
Reads product information and keywords from an Amazon Excel file.
Searches for corresponding products on Alibaba and stores 5 relevant results in a new Excel file.
"""

# Load the Excel file with Amazon product data
workbook = openpyxl.load_workbook('7.xlsx')
sheet = workbook.active

# Extract search queries, titles, and prices from the Excel sheet
search_queries = [row[6] for row in sheet.iter_rows(values_only=True, min_row=2)]
amazon_titles = [row[5] for row in sheet.iter_rows(values_only=True, min_row=2)]
amazon_prices = [float(row[0]) for row in sheet.iter_rows(values_only=True, min_row=2)]

# Set Firefox options to run in headless mode (without opening a visible browser)
options = Options()
options.add_argument('-headless')

# Initialize the Firefox WebDriver
driver = webdriver.Firefox(options=options)

# Create a new Excel workbook to store the results from Alibaba
result_workbook = openpyxl.Workbook()
result_sheet = result_workbook.active
result_sheet['A1'] = 'Search Query'
result_sheet['B1'] = 'Product Title'
result_sheet['C1'] = 'Image Link'
result_sheet['D1'] = 'Rating'
result_sheet['E1'] = 'Product Link'
result_sheet['F1'] = 'Keywords'
result_sheet['G1'] = 'Price'

# Open Alibaba homepage
driver.get("https://www.alibaba.com/")

# Start from the second row for storing results in the new Excel file
current_row = 2

# Loop through Amazon data to search for matching products on Alibaba
for amazon_title, search_query, amazon_price in zip(amazon_titles, search_queries, amazon_prices):
    product_count = 0  # Counter for the number of products per search query
    
    # Perform search on Alibaba
    for _ in range(5):
        try:
            # Locate and interact with the search bar on Alibaba
            search_bar = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "search-bar-input")))
            search_bar.clear()
            search_bar.send_keys(search_query)
            search_bar.send_keys(Keys.RETURN)
            time.sleep(5)  # Allow page to load
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '.elements-title-normal')))
            break
        except StaleElementReferenceException:
            continue  # Retry if there's a stale element exception

    # Parse the search results page with BeautifulSoup
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    products = soup.select('.elements-title-normal')

    # Extract product data from the page
    for product in products:
        if product_count >= 5:  # Limit to 5 products per query
            break
        
        # Extract title, link, and image
        title = product.text.strip()
        link = product['href']
        if link.startswith("//"):
            link = "https:" + link
        
        img = product.find_previous('img')
        img_link = 'No image found'
        if img and 'src' in img.attrs:
            img_link = img['src']
            if img_link.startswith('//'):
                img_link = 'https:' + img_link
            elif img_link.startswith('/'):
                img_link = 'https://www.alibaba.com' + img_link

        # Extract rating
        rating_element = soup.select_one('i.ui2-icon.ui2-icon-favorites-filling')
        rating = rating_element.next_sibling.strip() if rating_element and rating_element.next_sibling else None

        # Extract price and check if it's lower than the Amazon price
        price_element = product.find_previous('span', class_='elements-offer-price-normal__price')
        price = price_element.text.strip() if price_element else 'N/A'
        if price == 'N/A':
            continue
        
        # Convert price to float and compare with Amazon price
        prices = price.replace('US$', '').replace('€', '').replace(',', '').split('-')
        max_price = max(float(p) for p in prices)
        if max_price > amazon_price:
            continue

        # Write data to the result Excel file
        result_sheet.cell(row=current_row, column=1).value = amazon_title
        result_sheet.cell(row=current_row, column=2).value = title
        result_sheet.cell(row=current_row, column=3).value = img_link
        result_sheet.cell(row=current_row, column=4).value = rating
        result_sheet.cell(row=current_row, column=5).value = link
        result_sheet.cell(row=current_row, column=6).value = search_query
        result_sheet.cell(row=current_row, column=7).value = price

        # Save the result after every entry
        result_workbook.save('alibaba.xlsx')

        current_row += 1
        product_count += 1  # Increment the product counter

# Close the WebDriver
driver.quit()
