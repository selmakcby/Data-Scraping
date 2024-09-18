import openpyxl

""""
Combines two excel file. 
"""

# Load the 'results11.xlsx' file
results_workbook = openpyxl.load_workbook('alibaba1.xlsx')
results_sheet = results_workbook.active

# Load the existing Excel file
existing_workbook = openpyxl.load_workbook('xxxx.xlsx')
existing_sheet = existing_workbook.active

# Define the number of products per group
products_per_group = 5

# Calculate the number of groups
num_groups = (results_sheet.max_row - 1) // products_per_group  # Subtract 1 from max_row to skip the first row

# Copy the data for each group
for group in range(num_groups):
    start_row = group * products_per_group + 2  # Add 2 to start from the second row
    existing_row = group + 2  # Row index in the existing file

    for product_index in range(products_per_group):
        for col_index, col_letter in enumerate(['B', 'C', 'D', 'E', 'G'], start=1):
            results_col = ord(col_letter) - ord('A') + 1
            cell_value = results_sheet.cell(row=start_row + product_index, column=results_col).value
            target_col = (product_index * 5) + col_index + 6  # Adjust target column based on product index
            existing_sheet.cell(row=existing_row, column=target_col).value = f" {cell_value}"

# Save the changes back to the existing file
existing_workbook.save('xxxx.xlsx')
