import openpyxl

""""
Edits the excel file from alibaba_scaping.

"""
# Load the Excel file
workbook = openpyxl.load_workbook('alibaba.xlsx')
result_sheet = workbook.active

# Shift cells up in column G
for i in range(2, result_sheet.max_row):
    result_sheet.cell(row=i, column=7).value = result_sheet.cell(row=i+1, column=7).value

# Delete the last row
result_sheet.delete_rows(result_sheet.max_row)

# Save the changes
workbook.save('alibaba.xlsx')
